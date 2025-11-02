"""
Service de base de connaissances SCF
Charge et indexe le fichier SCF Excel pour permettre des recherches sémantiques précises
Avec cache persistant pour reprise après interruption
Utilise un modèle ML singleton partagé pour économiser la mémoire
"""

import openpyxl
from sentence_transformers import SentenceTransformer
import numpy as np
from typing import List, Dict, Optional, Tuple
import logging
import pickle
import os
import hashlib
from ml_model_singleton import get_shared_ml_model, get_model_name
from cache_config import CacheConfig

logger = logging.getLogger(__name__)

class SCFKnowledgeBase:
    """
    Base de connaissances SCF chargée depuis le fichier Excel
    Utilise des embeddings pour la recherche sémantique
    Utilise un modèle ML singleton partagé et cache centralisé
    """

    def __init__(self, excel_path: str = '/app/scf_knowledge_base.xlsx', cache_dir: Optional[str] = None):
        self.excel_path = excel_path

        # Utiliser la configuration centralisée du cache
        if cache_dir is None:
            self.cache_dir = str(CacheConfig.get_cache_dir())
        else:
            self.cache_dir = cache_dir
            # Créer le répertoire si spécifié manuellement
            os.makedirs(cache_dir, exist_ok=True)

        self.controls: List[Dict] = []
        self.threats: List[str] = []
        self.risks: List[str] = []

        # Modèle de similarité sémantique (partagé via singleton)
        self.model = None
        self.control_embeddings = None

        logger.info("📚 Initialisation de la base de connaissances SCF...")
        self._load_scf_controls()
        self._load_threats_and_risks()
        logger.info(f"✅ Base SCF chargée: {len(self.controls)} contrôles, {len(self.threats)} menaces, {len(self.risks)} risques")

    def _load_scf_controls(self):
        """Charge tous les contrôles SCF depuis la feuille 'SCF 2025.2'"""
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
            ws = wb['SCF 2025.2']

            logger.info(f"📖 Lecture de {ws.max_row} lignes de contrôles SCF...")

            # Lecture des contrôles (en sautant la ligne d'en-tête)
            for row_idx in range(2, ws.max_row + 1):
                scf_num = ws.cell(row_idx, 4).value  # Col 4: SCF #

                if not scf_num:  # Ligne vide
                    continue

                control = {
                    'scf_id': str(scf_num).strip(),
                    'scf_control': ws.cell(row_idx, 3).value,  # Col 3: SCF Control
                    'scf_domain': ws.cell(row_idx, 2).value,   # Col 2: SCF Domain
                    'description': ws.cell(row_idx, 5).value,  # Col 5: Control Description
                    'cobit_2019': ws.cell(row_idx, 6).value,   # Col 6: COBIT 2019
                    'control_question': ws.cell(row_idx, 13).value,  # Col 13: Control Question
                    'possible_solutions': ws.cell(row_idx, 10).value,  # Col 10: Medium Business Solutions
                }

                # Nettoyer les valeurs None
                for key in control:
                    if control[key] is None:
                        control[key] = ""
                    else:
                        control[key] = str(control[key]).strip()

                self.controls.append(control)

            wb.close()
            logger.info(f"✅ {len(self.controls)} contrôles SCF chargés")

        except Exception as e:
            logger.error(f"❌ Erreur lors du chargement des contrôles SCF: {e}")
            raise

    def _load_threats_and_risks(self):
        """Charge les catalogues de menaces et risques"""
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)

            # Threats
            if 'Threat Catalog' in wb.sheetnames:
                ws_threats = wb['Threat Catalog']
                for row_idx in range(2, ws_threats.max_row + 1):
                    threat = ws_threats.cell(row_idx, 1).value
                    if threat and str(threat).strip():
                        self.threats.append(str(threat).strip())

            # Risks
            if 'Risk Catalog' in wb.sheetnames:
                ws_risks = wb['Risk Catalog']
                for row_idx in range(2, ws_risks.max_row + 1):
                    risk = ws_risks.cell(row_idx, 1).value
                    if risk and str(risk).strip():
                        self.risks.append(str(risk).strip())

            wb.close()
            logger.info(f"✅ Catalogues chargés: {len(self.threats)} menaces, {len(self.risks)} risques")

        except Exception as e:
            logger.warning(f"⚠️ Erreur lors du chargement des catalogues: {e}")

    def _get_cache_key(self, model_name: str) -> str:
        """Génère une clé de cache unique basée sur le modèle et les contrôles"""
        # Hash des IDs de contrôles pour détecter les changements
        controls_hash = hashlib.md5(
            ''.join([ctrl['scf_id'] for ctrl in self.controls]).encode()
        ).hexdigest()
        return f"scf_embeddings_{model_name.replace('/', '_')}_{controls_hash}"

    def _get_cache_path(self, cache_key: str) -> str:
        """Retourne le chemin complet du fichier cache (format NumPy sécurisé)"""
        return os.path.join(self.cache_dir, f"{cache_key}.npz")

    def init_semantic_model(self, model: Optional[SentenceTransformer] = None):
        """
        Initialise le modèle de similarité sémantique et pré-calcule les embeddings
        Avec système de cache pour reprise après interruption

        Args:
            model: Modèle optionnel (si None, utilise le singleton partagé)
        """
        logger.info("🧠 Initialisation du modèle sémantique pour la base SCF...")

        # Utiliser le singleton si aucun modèle n'est fourni
        if model is None:
            logger.info("🔗 Utilisation du modèle ML partagé (singleton)")
            self.model = get_shared_ml_model()
            model_name = get_model_name()
        else:
            logger.info("⚠️ Utilisation d'un modèle ML custom (non recommandé)")
            self.model = model
            model_name = getattr(model, 'model_name', 'unknown_model')

        # Générer la clé de cache
        cache_key = self._get_cache_key(model_name)
        cache_path = self._get_cache_path(cache_key)

        # Vérifier si un cache existe
        if os.path.exists(cache_path):
            try:
                logger.info(f"💾 Cache trouvé: {cache_key}")
                # Charger depuis NumPy (SÉCURISÉ - pas pickle)
                cache_data = np.load(cache_path, allow_pickle=False)

                self.control_embeddings = cache_data['embeddings']
                cached_model = str(cache_data['model_name'][0])
                created_at = str(cache_data['created_at'][0])

                logger.info(f"✅ Embeddings chargés depuis le cache ({len(self.control_embeddings)} contrôles)")
                logger.info(f"📊 Modèle: {cached_model}, Date: {created_at}")
                return
            except Exception as e:
                logger.warning(f"⚠️ Erreur lecture cache NumPy: {e}")
                # Tentative de migration depuis ancien format pickle
                old_cache_path = cache_path.replace('.npz', '.pkl')
                if os.path.exists(old_cache_path):
                    logger.warning("💡 Tentative migration depuis ancien format pickle...")
                    try:
                        import pickle
                        with open(old_cache_path, 'rb') as f:
                            old_cache = pickle.load(f)
                        self.control_embeddings = old_cache['embeddings']
                        logger.info(f"✅ Migration réussie, {len(self.control_embeddings)} embeddings récupérés")
                        # Le nouveau cache sera créé automatiquement ci-dessous
                    except Exception as migration_error:
                        logger.error(f"❌ Migration impossible: {migration_error}")
                else:
                    logger.warning("⚠️ Recalcul nécessaire")

        # Pas de cache ou cache invalide -> calculer les embeddings
        logger.info("🔄 Aucun cache valide, calcul des embeddings...")

        # Créer des textes combinés pour chaque contrôle
        control_texts = []
        for ctrl in self.controls:
            combined_text = f"{ctrl['scf_id']} {ctrl['scf_control']} {ctrl['description']} {ctrl['control_question']}"
            control_texts.append(combined_text)

        # Calculer les embeddings par batch pour pouvoir reprendre
        logger.info(f"📊 Calcul des embeddings pour {len(control_texts)} contrôles...")
        logger.info("⏳ Cette opération peut prendre plusieurs minutes...")

        try:
            self.control_embeddings = self.model.encode(
                control_texts,
                show_progress_bar=False,
                batch_size=32  # Traiter par batch de 32
            )

            # Sauvegarder dans le cache au format NumPy (SÉCURISÉ)
            logger.info("💾 Sauvegarde des embeddings dans le cache (NumPy)...")
            from datetime import datetime

            np.savez_compressed(
                cache_path,
                embeddings=self.control_embeddings,
                model_name=np.array([model_name], dtype=object),
                num_controls=np.array([len(self.controls)], dtype=np.int32),
                created_at=np.array([datetime.now().isoformat()], dtype=object)
            )

            logger.info(f"✅ Embeddings calculés et sauvegardés (NumPy): {cache_path}")

            # Supprimer l'ancien cache pickle si présent
            old_cache_path = cache_path.replace('.npz', '.pkl')
            if os.path.exists(old_cache_path):
                try:
                    os.remove(old_cache_path)
                    logger.info("🗑️ Ancien cache pickle supprimé")
                except Exception:
                    pass

        except Exception as e:
            logger.error(f"❌ Erreur lors du calcul des embeddings: {e}")
            raise

    def find_best_scf_control(
        self,
        requirement_text: str,
        top_k: int = 5,
        min_similarity: float = 0.5
    ) -> List[Dict]:
        """
        Trouve les contrôles SCF les plus pertinents pour une exigence donnée

        Args:
            requirement_text: Texte de l'exigence à analyser
            top_k: Nombre de résultats à retourner
            min_similarity: Score de similarité minimum (0-1)

        Returns:
            Liste de contrôles SCF avec leurs scores de similarité
        """
        if not self.model or self.control_embeddings is None:
            raise RuntimeError("Le modèle sémantique n'est pas initialisé. Appelez init_semantic_model() d'abord.")

        # Encoder l'exigence
        req_embedding = self.model.encode([requirement_text])[0]

        # Calculer les similarités cosinus
        from sklearn.metrics.pairwise import cosine_similarity
        similarities = cosine_similarity([req_embedding], self.control_embeddings)[0]

        # Trier et récupérer les top_k
        top_indices = np.argsort(similarities)[::-1][:top_k]

        results = []
        for idx in top_indices:
            score = float(similarities[idx])
            if score >= min_similarity:
                control = self.controls[idx].copy()
                control['similarity_score'] = score
                results.append(control)

        return results

    def get_control_by_id(self, scf_id: str) -> Optional[Dict]:
        """Récupère un contrôle SCF par son ID exact"""
        scf_id_clean = scf_id.strip().upper()
        for ctrl in self.controls:
            if ctrl['scf_id'].upper() == scf_id_clean:
                return ctrl
        return None

    def find_related_cobit(self, scf_id: str) -> List[str]:
        """Trouve les références COBIT associées à un contrôle SCF"""
        ctrl = self.get_control_by_id(scf_id)
        if not ctrl or not ctrl['cobit_2019']:
            return []

        # Parser les références COBIT (séparées par des retours à la ligne)
        cobit_refs = [ref.strip() for ref in ctrl['cobit_2019'].split('\n') if ref.strip()]
        return cobit_refs

    def find_relevant_threat(self, requirement_text: str) -> Optional[str]:
        """Trouve la menace la plus pertinente depuis le catalogue"""
        if not self.threats or not self.model:
            return None

        # Encoder l'exigence et les menaces
        req_embedding = self.model.encode([requirement_text])[0]
        threat_embeddings = self.model.encode(self.threats, show_progress_bar=False)

        # Calculer similarités
        from sklearn.metrics.pairwise import cosine_similarity
        similarities = cosine_similarity([req_embedding], threat_embeddings)[0]

        # Prendre la menace la plus similaire si score > 0.3
        best_idx = np.argmax(similarities)
        if similarities[best_idx] > 0.3:
            return self.threats[best_idx]

        return None

    def find_relevant_risk(self, requirement_text: str) -> Optional[str]:
        """Trouve le risque le plus pertinent depuis le catalogue"""
        if not self.risks or not self.model:
            return None

        # Encoder l'exigence et les risques
        req_embedding = self.model.encode([requirement_text])[0]
        risk_embeddings = self.model.encode(self.risks, show_progress_bar=False)

        # Calculer similarités
        from sklearn.metrics.pairwise import cosine_similarity
        similarities = cosine_similarity([req_embedding], risk_embeddings)[0]

        # Prendre le risque le plus similaire si score > 0.3
        best_idx = np.argmax(similarities)
        if similarities[best_idx] > 0.3:
            return self.risks[best_idx]

        return None

    def validate_scf_reference(self, scf_ref: str) -> Tuple[bool, Optional[Dict]]:
        """
        Valide qu'une référence SCF existe réellement dans la base

        Returns:
            (is_valid, control_data)
        """
        # Extraire juste l'ID (ex: "SCF-IAC-02" ou "IAC-02")
        scf_id = scf_ref.split('-')[-2] + '-' + scf_ref.split('-')[-1] if '-' in scf_ref else scf_ref
        scf_id = scf_id.strip().upper()

        ctrl = self.get_control_by_id(scf_id)
        return (ctrl is not None, ctrl)


# Instance globale (singleton)
_scf_kb_instance: Optional[SCFKnowledgeBase] = None

def get_scf_knowledge_base() -> SCFKnowledgeBase:
    """Récupère l'instance singleton de la base de connaissances SCF"""
    global _scf_kb_instance
    if _scf_kb_instance is None:
        _scf_kb_instance = SCFKnowledgeBase()
    return _scf_kb_instance
